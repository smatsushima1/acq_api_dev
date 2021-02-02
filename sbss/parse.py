# this file parses the data from the size standards worksheet within the SBA Size
#   Standards spreadsheet into a json format
# footnotes are still prevalent within the data unfortunately...

import json
import openpyxl as opxl

###############################################################################
# functions

# parse-out naics and exceptions
def naicsParse(txt):
  estr_search = '_Except'
  if txt == 'None':
    nstr = 'N/A'
    estr = 'N/A'
  elif estr_search not in txt:
    nstr = txt
    estr = 'N/A'
  else:
    estr_s = txt.split('_')
    estr_c = int(len(estr_s))
    nstr = estr_s[0]
    if estr_c != 2:
      estr = 'Exception ' + estr_s[2]
    else:
      estr = 'Exception'
  return [nstr, estr]

# parse-out the subsector name
def subName(txt):
  subn_s = txt.split()
  subn_c = int(len(subn_s))
  subn = subn_s[3 : subn_c]
  subn = ' '.join(subn)
  return subn

# convert 'N/A' to numbers
# one of the values for ss_m is a string; this parses out the number from that
# otherwise, save it as a number as it should be
def convNA(txt):
  if txt.count(' ') > 0:
    ntxt = txt.split()
    ntxt = ntxt[0]
    ntxt = ntxt.replace('$', '')
    txt = ntxt
  elif txt == 'N/A':
    txt = 0
  return txt
###############################################################################

# load the wb
xl_file = 'SBA Table of Size Standards_Effective Aug 19, 2019.xlsx'
wb = opxl.load_workbook(xl_file) 

# find desired sheet name
# shts = wb.sheetnames
# for sname in shts:
#   print(sname)

# find max row and and column of sheet
ws = wb['table_of_size_standards-all']
mrow = ws.max_row
mcol = ws.max_column

# delete file if still in directory
# open file to be appended to
# add left bracket to signify the beginning of adding data
fname = 'sbss.json'
open(fname, 'w').close()
json_file = open(fname, 'a', encoding = 'utf8')
json_file.write('[')

# iterate through all rows
irows = ws.iter_rows(min_row = 3, max_row = mrow, max_col = mcol, values_only = True)
for r in irows:
  # create a new list to convert everything to strings, and change 'None' to 'N/A'
  r2 = []
  for i, val in enumerate(r):
    fval = str(r[i])
    
    # first start with naics value
    # the function will either save the value as a string, or parse out the
    #   exception to the naics
    if i == 0:
      nai_exc = naicsParse(fval)
      for j in nai_exc:
        r2.append(j)
    # if the values are emptoy, then change to 'N/A' - I don't want null values
    # otherwise, convert to string for now, and later change type accordingly
    elif fval == 'None':
      r2.append('N/A')
    else:
      r2.append(str(val))

  # work with new list
  # strip-out spaces to allow parsing
  # save subsector names to be used in function
  # continue to next iteration if the row doesn't apply in either cases
  fcol = r2[0]
  fcol = fcol.strip()
  if fcol[0:3] == 'Sub':
    subn = subName(r2[0])
    continue
  elif fcol == 'N/A':
    continue
  
  # begin assigning variables
  nai = r2[0]
  exc = r2[1]
  ssec = nai[0:3]
  ssec_n = subn
  desc = r2[2]
  ss_m = convNA(r2[3])
  ss_e = convNA(r2[4])
  fn = r2[5]
  
  # parse out the footnote
  if fn != 'N/A':
    fn_sp = fn.split()
    fn_co = int(len(fn_sp)) - 1
    fn = fn_sp[fn_co]
  
  # finally, pass all the data to a dictionary then dump into the file
  d = {
        'naics': int(nai),
        'exception': exc,
        'subsector': int(ssec),
        'subsector_name': ssec_n,
        'description': desc,
        'ss_millions': float(ss_m),
        'ss_employees': int(ss_e),
        'footnotes': fn
      }
  json.dump(d, json_file, indent = 2)

# add closing bracket to signify the end
json_file.write(']')
json_file.close()

# add commas inbetween each object
json_file = open(fname, 'r')
contents = json_file.read()
contents = contents.replace('}{', '},{')
json_file.close()

# overwrite old file with new changes
json_file = open(fname, 'w')
json_file.write(contents)

print('Finished pushing data')



