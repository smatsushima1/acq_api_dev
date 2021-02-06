# Parses the data from the size standards worksheet 
# Footnotes are still prevalent within the data unfortunately...

import os
from os import path
import json
import openpyxl as opxl

###############################################################################
# Functions

# Parse-out naics and exceptions
def naics_parse(txt):
  estr_search = '_Except'
  if txt == 'None':
    nstr = 'N/A'
    estr = 'N/A'
  elif estr_search not in txt:
    nstr = txt
    estr = 'N/A'
  else:
    estr_s = txt.split('_')
    estr_c = int(len(estr_s))
    nstr = estr_s[0]
    if estr_c != 2:
      estr = 'Exception ' + estr_s[2]
    else:
      estr = 'Exception'
  return [nstr, estr]

# Parse-out the subsector name
def sub_name(txt):
  subn_s = txt.split()
  subn_c = int(len(subn_s))
  subn = subn_s[3 : subn_c]
  subn = ' '.join(subn)
  return subn

# Convert 'N/A' to numbers
# One of the values for ss_m is a string; this parses out the number from that
# Otherwise, save it as a number as it should be
def conv_na(txt):
  if txt.count(' ') > 0:
    ntxt = txt.split()
    ntxt = ntxt[0]
    ntxt = ntxt.replace('$', '')
    txt = ntxt
  elif txt == 'N/A':
    txt = 0
  return txt
###############################################################################

# Load the wb
xl_file = 'SBA Table of Size Standards_Effective Aug 19, 2019.xlsx'
wb = opxl.load_workbook(xl_file) 

# Find desired sheet name
# shts = wb.sheetnames
# for sname in shts:
#   print(sname)

# Find max row and and column of sheet
ws = wb['table_of_size_standards-all']
mrow = ws.max_row
mcol = ws.max_column

# Load file
# Remove all its contents before writing anything, but only if it exists
jname = os.path.basename(__file__).split('.')[0]
jname = 'json/' + jname + '.json'

if path.exists(jname):
  open(jname, 'w').close()
  
with open(jname, 'w', encoding = 'utf8') as jf:
  jf.write('[')

  # Loop through all rows
  irows = ws.iter_rows(min_row = 3,
                       max_row = mrow,
                       max_col = mcol,
                       values_only = True
                       )
  for r in irows:
    # Create a new list to convert everything to strings, and change 'None' to 'N/A'
    r2 = []
    for i, val in enumerate(r):
      fval = str(r[i])
      
      # First start with naics value
      # The function will either save the value as a string, or parse out the
      #   exception to the naics
      if i == 0:
        nai_exc = naics_parse(fval)
        for j in nai_exc:
          r2.append(j)
      # If the values are emptoy, then change to 'N/A' - I don't want null values
      # Otherwise, convert to string for now, and later change type accordingly
      elif fval == 'None':
        r2.append('N/A')
      else:
        r2.append(str(val))
  
    # Work with new list
    # Strip-out spaces to allow parsing
    # Save subsector names to be used in function
    # Continue to next iteration if the row doesn't apply in either cases
    fcol = r2[0]
    fcol = fcol.strip()
    if fcol.startswith('Sub'):
      subn = sub_name(r2[0])
      continue
    elif fcol == 'N/A':
      continue
    
    # Begin assigning variables
    nai = r2[0]
    exc = r2[1]
    ssec = nai[0:3]
    ssec_n = subn
    desc = r2[2]
    ss_m = conv_na(r2[3])
    ss_e = conv_na(r2[4])
    fn = r2[5]
    
    # Parse out the footnote
    if fn != 'N/A':
      fn_sp = fn.split()
      fn_co = int(len(fn_sp)) - 1
      fn = fn_sp[fn_co]
    
    # Finally, pass all the data to a dictionary then dump into the file
    d = {
          'naics': int(nai),
          'exception': exc,
          'subsector': int(ssec),
          'subsector_name': ssec_n,
          'description': desc,
          'ss_millions': float(ss_m),
          'ss_employees': int(ss_e),
          'footnotes': fn
        }
    json.dump(d, jf, indent = 2)
  
  # Add closing bracket to signify the end
  jf.write(']')

# Add commas inbetween each object
with open(jname, 'r') as jf:
  contents = jf.read()
  contents = contents.replace('}{', '},{')

# Overwrite old file with new changes
with open(jname, 'w') as jf:
  jf.write(contents)

print('Finished pushing data to ' + jname)



