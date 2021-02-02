import openpyxl as opxl

xlfile = "vets4212.xlsx"
wb = opxl.load_workbook(xlfile)

# find desired sheet name
shts = wb.sheetnames
print(len(shts))

# for sname in shts:
#   print(sname)
  # print(sname.max_row)
  # print(sname.max_column)

# find max row and and column of sheet
# ws = wb["table_of_size_standards-all"]
# mrow = ws.max_row
# mcol = ws.max_column
# print(mrow)
# print(mcol)



















